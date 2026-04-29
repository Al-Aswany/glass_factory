"""Cutting Job controller — orchestrates the COP round-trip workflow."""
import json
from collections import defaultdict

import frappe
from frappe.model.document import Document
from frappe.utils import nowdate, nowtime
from openpyxl import Workbook, load_workbook

from glass_factory.glass_factory.cop_parser import (
	cross_validate,
	parse_label,
	parse_stock_diff,
	parse_tabular_files,
)
from glass_factory.glass_factory.stock_posting import build_stock_entries


class CuttingJob(Document):

	# ------------------------------------------------------------------
	# Lifecycle
	# ------------------------------------------------------------------

	def validate(self):
		"""Gate promotions to Result Uploaded so we never post bad data."""
		if self.status == "Result Uploaded":
			if not self.attached_stock_post:
				frappe.throw("Attach stock_post.xlsx before setting status to 'Result Uploaded'.")
			if not self.tabular_files:
				frappe.throw("At least one tabular file must be attached.")
			for tf in self.tabular_files:
				if not tf.attached_file or not tf.sheet_index:
					frappe.throw("Every tabular-file row must have a file and a sheet index.")

	# ------------------------------------------------------------------
	# Whitelisted actions (called from client-side buttons)
	# ------------------------------------------------------------------

	@frappe.whitelist()
	def pull_pieces_from_sales_orders(self):
		"""Populate the pieces child table from all linked SO cut_pieces rows."""
		self.pieces = []

		for so_link in self.linked_sales_orders or []:
			so_name = so_link.sales_order if hasattr(so_link, "sales_order") else so_link.get("sales_order")
			if not so_name:
				continue

			so = frappe.get_doc("Sales Order", so_name)
			for idx, cp in enumerate(so.cut_pieces or []):
				self.append("pieces", {
					"parent_item": cp.parent_item,
					"length_mm": cp.length_mm,
					"width_mm": cp.width_mm,
					"qty": cp.qty,
					"label": cp.user_label or f"Piece-{idx}",
					"customer_name": so.customer_name or so.customer,
					"sales_order": so_name,
					"sales_order_item": str(idx),  # 0-based index into cut_pieces
				})

		self.flags.ignore_validate = True
		self.save()
		return {"message": f"Pulled {len(self.pieces)} pieces from {len(self.linked_sales_orders or [])} SO(s)."}

	@frappe.whitelist()
	def generate_cop_files(self):
		"""Generate pieces.xlsx and stock_pre.xlsx for COP input."""
		if not self.pieces:
			frappe.throw("No pieces — pull from Sales Orders first.")
		if not self.source_sheets:
			frappe.throw("Select at least one source sheet.")

		# ---- pieces.xlsx ------------------------------------------------
		wb_pieces = Workbook()
		ws = wb_pieces.active
		ws.title = "Pieces"
		for idx, piece in enumerate(self.pieces):
			# Label format required by COP parser: "{user_label} | {SO}-{0-based-idx}"
			label = f"{piece.label} | {piece.sales_order}-{piece.sales_order_item}"
			ws.append([idx + 1, piece.length_mm, piece.width_mm, piece.qty,
					   piece.parent_item, 2, label, piece.customer_name])

		# ---- stock_pre.xlsx ---------------------------------------------
		wb_stock = Workbook()
		ws2 = wb_stock.active
		ws2.title = "Stock"
		for idx, (material, length, width, qty) in enumerate(self._aggregate_source_sheets()):
			ws2.append([idx + 1, length, width, qty, material, 2, material, 0])

		# ---- save and attach -------------------------------------------
		site_private = frappe.utils.get_site_path("private", "files")
		pieces_path = f"{site_private}/cop_pieces_{self.name}.xlsx"
		stock_path = f"{site_private}/cop_stock_pre_{self.name}.xlsx"
		wb_pieces.save(pieces_path)
		wb_stock.save(stock_path)

		self.attached_pieces = self._attach_file(pieces_path, f"cop_pieces_{self.name}.xlsx")
		self.attached_stock_pre = self._attach_file(stock_path, f"cop_stock_pre_{self.name}.xlsx")
		self.status = "Awaiting Optimization"
		self.flags.ignore_validate = True
		self.save()

		return {"message": "COP files generated.", "pieces": len(self.pieces)}

	@frappe.whitelist()
	def process_result(self):
		"""
		Parse the COP result files and return a confirmation payload.

		The payload is stored in self.flags so confirm_and_post() can access
		it in the same request without a round-trip.
		"""
		if self.status not in ("Awaiting Optimization", "Result Uploaded"):
			frappe.throw(f"Cannot process result from status '{self.status}'.")
		if not self.attached_stock_post:
			frappe.throw("Attach stock_post.xlsx first.")
		if not self.tabular_files:
			frappe.throw("Attach at least one tabular file.")

		pre_rows = self._load_excel_rows(self.attached_stock_pre)
		post_rows = self._load_excel_rows(self.attached_stock_post)
		consumed, remnants = parse_stock_diff(pre_rows, post_rows)

		tabular_data = [{"rows": self._load_excel_rows(tf.attached_file)} for tf in self.tabular_files]
		sheets = parse_tabular_files(tabular_data)

		warnings = cross_validate(consumed, sheets, [p.as_dict() for p in self.pieces])

		pieces_produced = sum(len(s["pieces"]) for s in sheets)
		scrap_m2 = self._compute_scrap_area(consumed, remnants, sheets)

		payload = {
			"consumed": [{"material": m, "length": l, "width": w, "qty": q} for m, l, w, q in consumed],
			"remnants": [{"material": m, "length": l, "width": w, "qty": q} for m, l, w, q in remnants],
			"sheets": sheets,
			"pieces_produced": pieces_produced,
			"remnants_created": len(remnants),
			"scrap_m2": scrap_m2,
			"warnings": warnings,
		}

		self.flags.parsed_payload = payload

		# Update summary fields and flip status
		self.pieces_produced = pieces_produced
		self.remnants_created = len(remnants)
		self.total_waste_m2 = scrap_m2
		consumed_area = sum((l * w / 1e6) * q for _, l, w, q in consumed)
		piece_area = sum(
			p["length"] * p["width"] / 1e6
			for s in sheets
			for p in s["pieces"]
		)
		self.utilization_pct = (piece_area / consumed_area * 100) if consumed_area else 0
		self.sheets_consumed = sum(int(q) for _, _l, _w, q in consumed)
		self.status = "Result Uploaded"
		self.flags.ignore_validate = True
		self.save()

		return payload

	@frappe.whitelist()
	def confirm_and_post(self, parsed_payload=None):
		"""
		Post Stock Entries and draft Delivery Notes.

		Status guard prevents duplicate posts from rapid double-clicks.
		"""
		if self.status != "Result Uploaded":
			frappe.throw(f"Cannot post from status '{self.status}'. Run 'Process Result' first.")

		if isinstance(parsed_payload, str):
			parsed_payload = json.loads(parsed_payload)
		if not parsed_payload:
			parsed_payload = self.flags.get("parsed_payload")
		if not parsed_payload:
			# Re-parse from the attached files (payload doesn't survive across requests)
			self.process_result()
			parsed_payload = self.flags.get("parsed_payload")

		# Build (do not insert yet) — fail fast before touching the DB
		stock_entries = build_stock_entries(self, parsed_payload)

		# Insert + submit each SE
		submitted_ses = []
		for se in stock_entries:
			se.insert(ignore_permissions=True)
			se.submit()
			submitted_ses.append(se.name)

		# Record the first SE for the dashboard link (one per spec, often just one)
		if submitted_ses:
			self.linked_stock_entry = submitted_ses[0]

		# Draft Delivery Notes — one per linked SO, for sales review
		dn_names = self._create_delivery_notes()
		self.linked_delivery_notes = ", ".join(dn_names)

		self.status = "Completed"
		self.flags.ignore_validate = True
		self.save()

		return {
			"message": "Posted successfully.",
			"stock_entries": submitted_ses,
			"delivery_notes": dn_names,
		}

	# ------------------------------------------------------------------
	# Private helpers
	# ------------------------------------------------------------------

	def _aggregate_source_sheets(self):
		"""
		Return a list of (material, length, width, qty) for the stock_pre file.

		Aggregates source_sheets by (item_code, length_mm, width_mm) — the
		same grouping COP uses to count available sheets.
		"""
		agg = defaultdict(float)
		for sheet in self.source_sheets:
			serial = frappe.get_doc("Serial No", sheet.serial_no)
			key = (sheet.item_code, int(serial.get("length_mm") or 0), int(serial.get("width_mm") or 0))
			agg[key] += 1

		# Flatten to list of 4-tuples
		return [(m, l, w, q) for (m, l, w), q in agg.items()]

	def _attach_file(self, file_path: str, filename: str) -> str:
		"""
		Attach a local file to this document using Frappe's file manager.

		Returns the file_url of the saved File document.
		"""
		with open(file_path, "rb") as fh:
			content = fh.read()

		file_doc = frappe.utils.file_manager.save_file(
			filename,
			content,
			"Cutting Job",
			self.name,
			is_private=1,
		)
		return file_doc.file_url

	def _load_excel_rows(self, file_url: str):
		"""
		Load rows from an attached Excel file.

		Uses the private-files path; file_url is the Frappe file URL
		(e.g. /private/files/foo.xlsx).
		"""
		filename = file_url.split("/")[-1]
		file_path = frappe.utils.get_site_path("private", "files", filename)
		wb = load_workbook(file_path, data_only=True)
		ws = wb.active

		rows = []
		headers = None
		for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
			if row_idx == 0:
				headers = [str(h) if h is not None else "" for h in row]
				continue
			if not any(row):
				break
			rows.append(dict(zip(headers, row)))

		return rows

	def _compute_scrap_area(self, consumed, remnants, sheets) -> float:
		consumed_area = sum((l * w / 1e6) * q for _, l, w, q in consumed)
		piece_area = sum(
			piece["length"] * piece["width"] / 1e6
			for sheet in sheets
			for piece in sheet["pieces"]
		)
		remnant_area = sum((l * w / 1e6) * q for _, l, w, q in remnants)
		return max(0.0, consumed_area - piece_area - remnant_area)

	def _create_delivery_notes(self):
		"""
		Create one draft Delivery Note per linked SO via ERPNext's own mapper.
		Returns a list of DN names.
		"""
		from erpnext.selling.doctype.sales_order.sales_order import make_delivery_note

		dn_names = []
		for so_link in self.linked_sales_orders or []:
			so_name = so_link.sales_order if hasattr(so_link, "sales_order") else so_link.get("sales_order")
			if not so_name:
				continue
			try:
				dn = make_delivery_note(so_name)
				dn.insert(ignore_permissions=True)
				dn_names.append(dn.name)
			except Exception:
				frappe.log_error(frappe.get_traceback(), f"DN creation failed for SO {so_name}")

		return dn_names
