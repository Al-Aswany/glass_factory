"""Layout Visualizer report for Cutting Job tabular layout rendering."""
import frappe
import json
import re
from collections import defaultdict
from typing import List, Dict, Tuple


PALETTE = [
	"#5B8FF9", "#5AD8A6", "#F6BD16", "#E86452", "#6DC8EC",
	"#945FB9", "#FF9845", "#1E9493", "#FF99C3", "#7666F9",
]


def execute(filters=None):
	"""Generate layout visualization for a Cutting Job."""
	if not filters or not filters.get("cutting_job"):
		return _empty_state("Select a Cutting Job to render its layout.")

	job_name = filters.get("cutting_job")
	if not frappe.db.exists("Cutting Job", job_name):
		return _empty_state(f"Cutting Job <b>{frappe.utils.escape_html(job_name)}</b> not found.")

	job = frappe.get_doc("Cutting Job", job_name)

	if not job.tabular_files:
		return _empty_state(
			"This Cutting Job has no tabular layout files attached yet. "
			"Run the optimizer or attach the Pieces XLSX outputs to render the layout."
		)

	sheets_data = _load_all_sheets(job)
	color_map = _build_color_map(sheets_data)

	html_blocks = [_styles(), _header(job)]
	for sheet_idx in sorted(sheets_data.keys()):
		pieces = sheets_data[sheet_idx]
		sheet_meta = _sheet_meta(job, sheet_idx)
		html_blocks.append(_render_sheet_card(sheet_idx, pieces, sheet_meta, color_map))
	html_blocks.append(_render_legend(color_map))

	message = "\n".join(html_blocks)

	columns, data = _build_table(sheets_data)
	report_summary = _build_summary(job, sheets_data)

	return columns, data, message, None, report_summary


def _empty_state(msg):
	html = f'<div class="lv-empty">{msg}</div>{_styles()}'
	return [], [], html, None, None


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_all_sheets(job) -> Dict[int, List[Dict]]:
	"""Load every tabular file and group rows by sheet index."""
	sheets_data: Dict[int, List[Dict]] = defaultdict(list)
	for tf in job.tabular_files:
		try:
			rows = _load_excel_rows(tf.attached_file)
		except Exception as e:  # noqa: BLE001
			frappe.log_error(f"Layout Visualizer: failed to load {tf.attached_file}: {e}")
			rows = []
		sheets_data[int(tf.sheet_index or 0)].extend(rows)
	return sheets_data


def _load_excel_rows(file_url: str) -> List[Dict]:
	from openpyxl import load_workbook

	file_name = file_url.split("/")[-1]
	file_path = frappe.get_site_path("private", "files", file_name)

	wb = load_workbook(file_path, data_only=True)
	ws = wb.active

	rows: List[Dict] = []
	headers = None
	for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
		if row_idx == 0:
			headers = [str(h or "").strip() for h in row]
			continue
		if not any(c is not None and c != "" for c in row):
			break
		rows.append(dict(zip(headers, row)))
	return rows


def _sheet_meta(job, sheet_idx: int) -> Dict:
	"""Return source-sheet metadata for a given sheet index (1-based)."""
	if not job.source_sheets:
		return {"length_mm": 0, "width_mm": 0, "item_code": "", "serial_no": ""}

	idx = max(1, min(sheet_idx, len(job.source_sheets))) - 1
	src = job.source_sheets[idx]
	return {
		"length_mm": float(src.length_mm or 0),
		"width_mm": float(src.width_mm or 0),
		"item_code": src.item_code or "",
		"serial_no": src.serial_no or "",
	}


def _piece_key(label: str) -> str:
	"""Group identical piece types into the same color (strip SO suffix)."""
	if not label:
		return "?"
	return re.split(r"\s*\|\s*", str(label))[0].strip()


def _build_color_map(sheets_data: Dict[int, List[Dict]]) -> Dict[str, str]:
	keys = []
	seen = set()
	for pieces in sheets_data.values():
		for p in pieces:
			k = _piece_key(p.get("Label", ""))
			if k not in seen:
				seen.add(k)
				keys.append(k)
	return {k: PALETTE[i % len(PALETTE)] for i, k in enumerate(keys)}


# ---------------------------------------------------------------------------
# HTML rendering
# ---------------------------------------------------------------------------

def _header(job) -> str:
	status = frappe.utils.escape_html(job.status or "")
	return f"""
	<div class="lv-header">
		<div class="lv-title">
			<div class="lv-job">{frappe.utils.escape_html(job.name)}</div>
			<div class="lv-sub">Cutting Layout Visualization</div>
		</div>
		<div class="lv-status lv-status-{status.lower()}">{status}</div>
	</div>
	"""


def _render_sheet_card(sheet_idx: int, pieces: List[Dict], meta: Dict, color_map: Dict[str, str]) -> str:
	if not pieces:
		return f'<div class="lv-card"><div class="lv-card-head">Sheet {sheet_idx}</div>'\
			f'<div class="lv-empty">No pieces placed on this sheet.</div></div>'

	sheet_w = meta["length_mm"] or _infer_sheet_extent(pieces, "x")
	sheet_h = meta["width_mm"] or _infer_sheet_extent(pieces, "y")

	# Render area: keep aspect ratio, max width 1100px
	max_w = 1100.0
	scale = max_w / sheet_w if sheet_w else 0.2
	pad = 24

	view_w = int(sheet_w * scale + 2 * pad)
	view_h = int(sheet_h * scale + 2 * pad)

	used_area = sum(float(p.get("Length", 0)) * float(p.get("Width", 0)) for p in pieces)
	sheet_area = sheet_w * sheet_h if sheet_w and sheet_h else 1
	utilization = (used_area / sheet_area) * 100 if sheet_area else 0

	svg_parts = [
		f'<svg viewBox="0 0 {view_w} {view_h}" class="lv-svg" xmlns="http://www.w3.org/2000/svg">',
		'<defs>',
		'  <pattern id="lv-grid" width="20" height="20" patternUnits="userSpaceOnUse">',
		'    <path d="M 20 0 L 0 0 0 20" fill="none" stroke="rgba(0,0,0,0.06)" stroke-width="1"/>',
		'  </pattern>',
		'  <filter id="lv-shadow" x="-10%" y="-10%" width="120%" height="120%">',
		'    <feDropShadow dx="0" dy="1" stdDeviation="1.2" flood-opacity="0.18"/>',
		'  </filter>',
		'</defs>',
		# Sheet background
		f'<rect x="{pad}" y="{pad}" width="{sheet_w * scale:.1f}" height="{sheet_h * scale:.1f}" '
		f'fill="url(#lv-grid)" stroke="#2c3e50" stroke-width="2" rx="2"/>',
	]

	for p in pieces:
		length = float(p.get("Length", 0))
		width = float(p.get("Width", 0))
		left = float(p.get("Left", 0))
		top = float(p.get("Top", 0))
		rotated = bool(p.get("Rotated", 0))
		label = str(p.get("Label", "") or "")
		key = _piece_key(label)
		color = color_map.get(key, "#888")

		x = pad + left * scale
		y = pad + top * scale
		w = length * scale
		h = width * scale

		short_label = key
		dim_label = f"{int(length)} × {int(width)}"
		tooltip = frappe.utils.escape_html(f"{label}  •  {dim_label} mm  •  ({int(left)},{int(top)})")

		svg_parts.append(
			f'<g class="lv-piece"><title>{tooltip}</title>'
			f'<rect x="{x:.1f}" y="{y:.1f}" width="{w:.1f}" height="{h:.1f}" '
			f'fill="{color}" fill-opacity="0.78" stroke="#1f2d3d" stroke-width="1" '
			f'filter="url(#lv-shadow)" rx="2"/>'
		)

		# Labels only if there is room
		if w > 60 and h > 30:
			cx = x + w / 2
			cy = y + h / 2
			svg_parts.append(
				f'<text x="{cx:.1f}" y="{cy - 4:.1f}" text-anchor="middle" '
				f'class="lv-piece-label">{frappe.utils.escape_html(short_label)}</text>'
			)
			svg_parts.append(
				f'<text x="{cx:.1f}" y="{cy + 10:.1f}" text-anchor="middle" '
				f'class="lv-piece-dim">{dim_label}</text>'
			)
			if rotated:
				svg_parts.append(
					f'<text x="{x + w - 6:.1f}" y="{y + 12:.1f}" text-anchor="end" '
					f'class="lv-piece-rot">↻</text>'
				)
		elif w > 30 and h > 14:
			svg_parts.append(
				f'<text x="{x + w/2:.1f}" y="{y + h/2 + 3:.1f}" text-anchor="middle" '
				f'class="lv-piece-dim">{frappe.utils.escape_html(short_label)}</text>'
			)

		svg_parts.append('</g>')

	svg_parts.append('</svg>')

	source_label = ""
	if meta["item_code"]:
		source_label = f' • {frappe.utils.escape_html(meta["item_code"])}'
		if meta["serial_no"]:
			source_label += f' #{frappe.utils.escape_html(meta["serial_no"])}'

	return f"""
	<div class="lv-card">
		<div class="lv-card-head">
			<div>
				<span class="lv-sheet-tag">Sheet {sheet_idx}</span>
				<span class="lv-sheet-dim">{int(sheet_w)} × {int(sheet_h)} mm{source_label}</span>
			</div>
			<div class="lv-metrics">
				<span class="lv-metric"><b>{len(pieces)}</b> pieces</span>
				<span class="lv-metric lv-util"><b>{utilization:.1f}%</b> utilization</span>
			</div>
		</div>
		<div class="lv-svg-wrap">{"".join(svg_parts)}</div>
	</div>
	"""


def _infer_sheet_extent(pieces: List[Dict], axis: str) -> float:
	if axis == "x":
		return max((float(p.get("Left", 0)) + float(p.get("Length", 0))) for p in pieces) or 1
	return max((float(p.get("Top", 0)) + float(p.get("Width", 0))) for p in pieces) or 1


def _render_legend(color_map: Dict[str, str]) -> str:
	if not color_map:
		return ""
	chips = "".join(
		f'<span class="lv-chip"><i style="background:{color}"></i>'
		f'{frappe.utils.escape_html(key)}</span>'
		for key, color in color_map.items()
	)
	return f'<div class="lv-legend"><span class="lv-legend-title">Legend</span>{chips}</div>'


# ---------------------------------------------------------------------------
# Tabular columns/data + summary
# ---------------------------------------------------------------------------

def _build_table(sheets_data: Dict[int, List[Dict]]) -> Tuple[List[Dict], List[Dict]]:
	columns = [
		{"fieldname": "sheet", "label": "Sheet", "fieldtype": "Int", "width": 80},
		{"fieldname": "piece", "label": "Piece", "fieldtype": "Data", "width": 220},
		{"fieldname": "length_mm", "label": "Length (mm)", "fieldtype": "Int", "width": 110},
		{"fieldname": "width_mm", "label": "Width (mm)", "fieldtype": "Int", "width": 110},
		{"fieldname": "left_mm", "label": "Left", "fieldtype": "Int", "width": 90},
		{"fieldname": "top_mm", "label": "Top", "fieldtype": "Int", "width": 90},
		{"fieldname": "rotated", "label": "Rotated", "fieldtype": "Check", "width": 80},
		{"fieldname": "customer", "label": "Customer", "fieldtype": "Data", "width": 200},
	]

	data = []
	for sheet_idx in sorted(sheets_data.keys()):
		for p in sheets_data[sheet_idx]:
			data.append({
				"sheet": sheet_idx,
				"piece": p.get("Label") or "",
				"length_mm": int(float(p.get("Length", 0) or 0)),
				"width_mm": int(float(p.get("Width", 0) or 0)),
				"left_mm": int(float(p.get("Left", 0) or 0)),
				"top_mm": int(float(p.get("Top", 0) or 0)),
				"rotated": 1 if p.get("Rotated") else 0,
				"customer": p.get("Customer name") or "",
			})
	return columns, data


def _build_summary(job, sheets_data: Dict[int, List[Dict]]) -> List[Dict]:
	total_pieces = sum(len(v) for v in sheets_data.values())
	return [
		{"label": "Sheets Consumed", "value": job.sheets_consumed or len(sheets_data),
		 "indicator": "Blue", "datatype": "Int"},
		{"label": "Pieces Produced", "value": job.pieces_produced or total_pieces,
		 "indicator": "Green", "datatype": "Int"},
		{"label": "Utilization", "value": f"{job.utilization_pct or 0:.2f}%",
		 "indicator": "Orange", "datatype": "Data"},
		{"label": "Remnants Created", "value": job.remnants_created or 0,
		 "indicator": "Grey", "datatype": "Int"},
		{"label": "Waste (m²)", "value": f"{job.total_waste_m2 or 0:.3f}",
		 "indicator": "Red", "datatype": "Data"},
	]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def _styles() -> str:
	return """
<style>
.lv-empty {
	padding: 28px; text-align: center; color: var(--text-muted);
	background: var(--bg-color); border: 1px dashed var(--border-color);
	border-radius: 8px; margin: 16px 0;
}
.lv-header {
	display:flex; justify-content:space-between; align-items:center;
	padding: 16px 20px; margin-bottom: 16px;
	background: linear-gradient(135deg, var(--bg-color) 0%, var(--fg-color) 100%);
	border: 1px solid var(--border-color); border-radius: 10px;
}
.lv-title .lv-job { font-size: 18px; font-weight: 700; color: var(--text-color); }
.lv-title .lv-sub { font-size: 12px; color: var(--text-muted); margin-top: 2px; }
.lv-status {
	padding: 6px 14px; border-radius: 999px; font-size: 12px; font-weight: 600;
	text-transform: uppercase; letter-spacing: 0.4px;
	background: rgba(46, 204, 113, 0.15); color: #2ecc71;
}
.lv-status-draft { background: rgba(149,165,166,0.15); color: #95a5a6; }
.lv-status-completed { background: rgba(46,204,113,0.15); color: #2ecc71; }
.lv-status-in.progress, .lv-status-in-progress { background: rgba(52,152,219,0.15); color: #3498db; }

.lv-card {
	background: var(--fg-color);
	border: 1px solid var(--border-color);
	border-radius: 10px;
	padding: 14px 16px 16px;
	margin-bottom: 18px;
	box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.lv-card-head {
	display:flex; justify-content:space-between; align-items:center;
	flex-wrap: wrap; gap: 10px; margin-bottom: 10px;
}
.lv-sheet-tag {
	display:inline-block; background: var(--primary); color:#fff;
	padding: 4px 10px; border-radius: 6px; font-weight: 600; font-size: 13px;
	margin-right: 10px;
}
.lv-sheet-dim { color: var(--text-muted); font-size: 13px; }
.lv-metrics { display:flex; gap: 14px; }
.lv-metric { font-size: 12px; color: var(--text-muted); }
.lv-metric b { color: var(--text-color); font-size: 14px; margin-right: 3px; }
.lv-util b { color: #2ecc71; }

.lv-svg-wrap { width: 100%; overflow-x: auto; background:
	repeating-linear-gradient(45deg, transparent, transparent 6px,
	rgba(127,127,127,0.04) 6px, rgba(127,127,127,0.04) 12px);
	padding: 10px; border-radius: 6px;
}
.lv-svg { width: 100%; height: auto; display: block; }
.lv-piece { transition: opacity 0.15s ease; cursor: pointer; }
.lv-piece:hover { opacity: 0.85; }
.lv-piece-label { font: 600 11px/1 Inter, system-ui, sans-serif; fill: #1f2d3d;
	pointer-events: none; }
.lv-piece-dim { font: 500 10px/1 Inter, system-ui, sans-serif; fill: #1f2d3d;
	opacity: 0.78; pointer-events: none; }
.lv-piece-rot { font: 700 12px/1 system-ui, sans-serif; fill: #c0392b;
	pointer-events: none; }

.lv-legend {
	display:flex; flex-wrap: wrap; gap: 10px; align-items:center;
	padding: 12px 14px; background: var(--fg-color);
	border: 1px solid var(--border-color); border-radius: 8px;
	margin: 4px 0 16px;
}
.lv-legend-title { font-weight: 600; font-size: 12px; color: var(--text-muted);
	text-transform: uppercase; letter-spacing: 0.5px; margin-right: 6px; }
.lv-chip {
	display:inline-flex; align-items:center; gap: 6px;
	padding: 4px 10px; background: var(--bg-color);
	border: 1px solid var(--border-color); border-radius: 999px;
	font-size: 12px; color: var(--text-color);
}
.lv-chip i { width: 10px; height: 10px; border-radius: 2px;
	display:inline-block; }

@media (max-width: 720px) {
	.lv-header { flex-direction: column; align-items: flex-start; gap: 10px; }
	.lv-card-head { flex-direction: column; align-items: flex-start; }
}
</style>
"""
